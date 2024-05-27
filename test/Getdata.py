import requests
import pandas as pd
import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def fetch_data(url, headers):
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()['data']


def process_data(data):
    df = pd.DataFrame(data)
    df.dropna(axis=1, how='all', inplace=True)
    return df


def clean_data(df):
    # 筛选代码以'6', '0', '3'开头的行
    filtered_df = df[df['code'].str.startswith(('6', '0', '3'))]

    # 排除名称中包含"国债"的行
    filtered_df = filtered_df[~filtered_df['name'].str.contains("债")]

    filtered_df.rename(
        columns={
            'name': '股票名称',
            'code': '股票代码',
            'date': '更新日期',
            'stkGroup': '集中分组'},
        inplace=True)

    # 格式化日期列为"月-日"格式
    filtered_df['更新日期'] = filtered_df['更新日期'].apply(
        lambda x: datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S').strftime('%m-%d'))
    return filtered_df


def write_data_to_excel(df, excel_path='StkGroupData.xlsx'):
    # 内部函数定义：检查是否满足标红的条件
    def should_highlight_red(stkGroup):
        return stkGroup.value is None or stkGroup.value == '' or stkGroup.value == 'E'

    # 内部函数定义：检查是否满足标黄的条件
    def should_highlight_yellow(fro, stkGroup):
        # 融资比例为空或大于1
        financing_condition = fro.value is None or fro.value == '' or (
                isinstance(fro.value, (int, float)) and fro.value > 1)
        # 集中分组为D
        group_condition = stkGroup.value and stkGroup.value == 'D'
        return financing_condition or group_condition

    # 更新顺序
    new_order = ['股票代码', '股票名称', '融资比例', '冲抵比例', '集中分组', '最新价', '流通市值', '总市值',
                 '成交额', '涨跌幅', '更新日期']
    df = df.reindex(columns=new_order)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        font = Font(name='黑体', size=16)
        alignment = Alignment(horizontal='center', vertical='center')

        # 定义标红和标黄的填充样式
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            # 获取相关列的单元格
            group_cell = row[new_order.index('集中分组')]
            financing_ratio_cell = row[new_order.index('融资比例')]

            # 检查是否满足 标红条件
            if should_highlight_red(group_cell):
                for cell in row:
                    cell.fill = red_fill
            # 检查是否满足标黄条件
            elif should_highlight_yellow(financing_ratio_cell, group_cell):
                for cell in row:
                    cell.fill = yellow_fill

            # 应用字体和对齐样式
            for cell in row:
                cell.font = font
                cell.alignment = alignment

        # 设置第一行加粗，字体大小为18
        for cell in worksheet["1:1"]:
            cell.font = Font(name='黑体', bold=True, size=18)

        # 设置列宽，为16个字符
        for column_cells in worksheet.columns:
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = 16

        # 设置行高，为25磅
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 25


def get_financial_data():
    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Mobile Safari/537.36",
        "Referer": "http://www.chinastock.com.cn/newsite/cgs-services/stockFinance/businessAnnc.html?type=marginList",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Cookie": "sensorsdata2015jssdkcross=...; aliyungf_tc=...; acw_tc=..."
    }

    url_template = "http://www.chinastock.com.cn/website2020/margin/stockList?dataType={}&stockCode=&queryDate="
    data_types = {'1': '融资比例', '3': '冲抵比例'}

    dfs = []
    for data_type, column_name in data_types.items():
        url = url_template.format(data_type)
        data = fetch_data(url, headers)
        df = process_data(data)
        df.rename(columns={'rate': column_name}, inplace=True)
        dfs.append(df)

    # 以联合主键合并表格
    merged_df = pd.merge(dfs[0], dfs[1], on=['code', 'name', 'date', 'stkGroup'], how='outer')
    # 进行数据清洗
    cleaned_df = clean_data(merged_df)
    # 获取总表
    allstocklist = get_all_stock_list()
    allstocklist['股票代码'] = allstocklist['股票代码'].astype(str).str.zfill(6).str.strip()
    cleaned_df['股票代码'] = cleaned_df['股票代码'].astype(str).str.zfill(6).str.strip()
    # 假设df1和df2是要合并的两个DataFrame，我们只想保留df1中的"股票名称"
    merged_df = pd.merge(allstocklist, cleaned_df, on='股票代码', how='left', suffixes=('', '_drop'))
    merged_df = merged_df.drop([col for col in merged_df.columns if 'drop' in col], axis=1)

    # 调用函数进行数据写入
    write_data_to_excel(merged_df)


def get_all_stock_list():
    # 给定的URL
    url = ('http://11.push2.eastmoney.com/api/qt/clist/get?cb=jQuery112405762645058319873_1713882388617&pn=1&pz=6000'
           '&po=0&np=1&ut=bd1d9ddb04089700cf9c27f6f7426281&fltt=2&invt=2&wbp2u=|0|0|0|web&fid=f6&fs=m:0+t:6,m:0+t:80,'
           'm:1+t:2,m:1+t:23,m:0+t:81+s:2048&fields=f12,f14,f2,f3,f6,f20,f21')

    # 发送HTTP请求并获取响应文本
    response_text = requests.get(url).text

    # 提取JSON字符串，去除JSONP的函数调用部分
    json_str = response_text[response_text.index('(') + 1: -2]

    # 将字符串解析为JSON格式
    response_json = json.loads(json_str)

    # 从JSON数据中提取股票信息列表
    stocks_data = response_json["data"]["diff"]

    # 创建DataFrame
    df = pd.DataFrame(stocks_data)
    # 将表头从字段编码更改为对应的名字
    df.rename(columns={
        'f2': '最新价',
        'f3': '涨跌幅',
        'f6': '成交额',
        'f12': '股票代码',
        'f14': '股票名称',
        'f20': '总市值',
        'f21': '流通市值',
    }, inplace=True)
    # 将DataFrame倒序
    df = df.iloc[::-1].reset_index(drop=True)
    # 输出DataFrame
    return df


# 调用函数
get_financial_data()
