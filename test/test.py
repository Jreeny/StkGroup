import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 打开Excel文件
workbook = openpyxl.load_workbook('StkGroupData.xlsx')

# 选择工作表
worksheet = workbook.active
new_order = ['股票代码', '股票名称', '融资比例', '冲抵比例', '集中分组', '最新价', '流通市值', '总市值',
             '成交额', '涨跌幅', '更新日期']


def should_highlight_yellow(financing_ratio_cell, group_cell):
    # 融资比例为空或大于1
    financing_condition = financing_ratio_cell.value is None or (
            isinstance(financing_ratio_cell.value, (int, float)) and financing_ratio_cell.value > 1)
    # 集中分组为D
    group_condition = group_cell.value and group_cell.value == 'D'
    return financing_condition or group_condition


def should_highlight_red(group_cell):
    return group_cell.value is None


font = Font(name='黑体', size=16)
alignment = Alignment(horizontal='center', vertical='center')

# 定义标红和标黄的填充样式
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    # 获取相关列的单元格
    group_cell = row[new_order.index('集中分组')]
    financing_ratio_cell = row[new_order.index('融资比例')]

    # 检查是否满足 标红条件
    if should_highlight_red(group_cell):
        for cell in row:
            cell.fill = red_fill
            print("fuckme")
    # 检查是否满足标黄条件
    elif should_highlight_yellow(financing_ratio_cell, group_cell):
        for cell in row:
            cell.fill = yellow_fill

    # 应用字体和对齐样式
    for cell in row:
        cell.font = font
        cell.alignment = alignment

# 设置第一行加粗，字体大小为18
for cell in worksheet["1:1"]:
    cell.font = Font(name='黑体', bold=True, size=18)

# 设置列宽，为16个字符
for column_cells in worksheet.columns:
    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = 16

# 设置行高，为25磅
for row in worksheet.iter_rows():
    worksheet.row_dimensions[row[0].row].height = 25



# 关闭Excel文件
workbook.close()
